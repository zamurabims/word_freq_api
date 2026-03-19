import io
from typing import Dict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.domain.entities.word_frequency import WordFrequency


class ExcelReportBuilder:
    """Строит xlsx-отчёт по частотной статистике."""

    HEADER_FILL = PatternFill("solid", start_color="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    CELL_FONT = Font(name="Arial", size=10)
    BORDER_SIDE = Side(style="thin", color="BFBFBF")
    CELL_BORDER = Border(
        left=BORDER_SIDE, right=BORDER_SIDE,
        top=BORDER_SIDE, bottom=BORDER_SIDE,
    )
    ALT_FILL = PatternFill("solid", start_color="EBF3FB")

    HEADERS = ["Словоформа (лемма)", "Кол-во в документе", "Кол-во по строкам"]
    COL_WIDTHS = [25, 22, 60]

    def build(
        self,
        stats: Dict[str, WordFrequency],
        total_lines: int,
    ) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Частотная статистика"

        self._write_headers(ws)

        sorted_words = sorted(stats.values(), key=lambda w: w.total_count, reverse=True)
        for row_idx, wf in enumerate(sorted_words, start=2):
            line_str = wf.line_counts_as_string(total_lines)
            row_data = [wf.lemma, wf.total_count, line_str]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.CELL_FONT
                cell.border = self.CELL_BORDER
                cell.alignment = Alignment(
                    horizontal="left" if col_idx != 2 else "center",
                    vertical="center",
                    wrap_text=(col_idx == 3),
                )
                if row_idx % 2 == 0:
                    cell.fill = self.ALT_FILL

        self._set_column_widths(ws)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:C{len(sorted_words) + 1}"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    def _write_headers(self, ws) -> None:
        for col_idx, (header, width) in enumerate(
            zip(self.HEADERS, self.COL_WIDTHS), start=1
        ):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.CELL_BORDER
        ws.row_dimensions[1].height = 20

    def _set_column_widths(self, ws) -> None:
        for col_idx, width in enumerate(self.COL_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
